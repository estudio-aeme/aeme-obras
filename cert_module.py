"""
cert_module.py — Módulo de certificación para el bot ObraManager
Funcionalidad:
  - Descarga el xlsx del Drive
  - Lee el último cert y sus acumulados
  - Genera nueva pestaña CERT N°X con los avances nuevos
  - Genera PDF del certificado
  - Sube ambos archivos al Drive
  - Registra en la base de datos
"""

import os, io, json, base64, requests, tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy

# ─── IDs de archivos en Google Drive (INTERNO) ────────────────────────────────
DRIVE_FILES = {
    "durlock":     {"xlsx": "1oU_5BEzovYxkzkspgF3_NJvr_H71QXO8",
                    "folder": "1S7jMjkeILwL2CqG7GasEL8ksLRc7pu0F"},
    "electricidad": {"xlsx": "1_3nTIIu1FsCHYH9FG3vznip97GpPVy6w",
                     "folder": "1CFSIvx02W95erSUJAtrOPYxHwkvZxX2E"},
    "sanitarias":  {"xlsx": "1F5vFqXlC10FpjL9yylC3y474lwPkjh4S",
                    "folder": "1AufrPH_vzRURf7N6ks5yhqcfXWVEzJse"},
    "pre_aa":      {"xlsx": "1BhWZWJt4LDU1BU-P2hgpDDixkFxDppqd",
                    "folder": "1d7gODVbgNxw6UB06KdO_fXpp7JlNbSCk"},
    "herreria":    {"xlsx": "1ZV5Q9pQLnudPk8wA04uqD6qhVODOYO-4",
                    "folder": "1Tj20i2ry813fsWx3anDbV0afRgS0mQu5"},
}

# IDs carpetas CLIENTE para los certs de venta
DRIVE_FILES_CLIENTE = {
    "durlock":      {"folder": "14-uy2s03iepJj7-20RLXJSCXdv0qtPQf",
                     "xlsx": "15eGIY9ZsbzuK6bknfd7Nt32wtUxxg1pq"},
    "electricidad": {"folder": "1dU-8ND_lHCGFRbvD6wzVgw4WuYWWzWBF",
                     "xlsx": "179DiNZoDnCXdVbg06Wn1mDPdZ1Kl0rwR"},
    "sanitarias":   {"folder": "1s5qNo54ocGKmYMgMo3Tcs6Z7HhYPMkyS",
                     "xlsx": "1oFI357pHLDgGhl8Wi8klroSFnl86XtwU"},
    "pre_aa":       {"folder": "1d7gODVbgNxw6UB06KdO_fXpp7JlNbSCk",
                     "xlsx": "1BhWZWJt4LDU1BU-P2hgpDDixkFxDppqd"},
}

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")


def _get_access_token():
    """Obtener access token de Google via metadata server (Render)."""
    # En Render con Google OAuth configurado
    token = os.environ.get("GOOGLE_ACCESS_TOKEN")
    if token:
        return token
    # Fallback: service account via env
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if sa_json:
        import google.auth
        from google.oauth2 import service_account
        from google.auth.transport.requests import Request
        creds = service_account.Credentials.from_service_account_info(
            json.loads(sa_json),
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        creds.refresh(Request())
        return creds.token
    return None


def drive_download(file_id):
    """Descarga un archivo de Google Drive. Retorna bytes."""
    token = _get_access_token()
    if not token:
        raise Exception("No hay token de Google Drive configurado")
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    r.raise_for_status()
    return r.content


def drive_upload_new(folder_id, filename, content_bytes, mime_type):
    """Sube un archivo nuevo a una carpeta de Drive. Retorna el file_id."""
    token = _get_access_token()
    metadata = {"name": filename, "parents": [folder_id]}
    r = requests.post(
        "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "metadata": (None, json.dumps(metadata), "application/json"),
            "file": (filename, content_bytes, mime_type),
        }
    )
    r.raise_for_status()
    return r.json()["id"]


def drive_update(file_id, content_bytes, mime_type):
    """Actualiza el contenido de un archivo existente en Drive."""
    token = _get_access_token()
    r = requests.patch(
        f"https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media",
        headers={"Authorization": f"Bearer {token}", "Content-Type": mime_type},
        data=content_bytes
    )
    r.raise_for_status()
    return r.json()


# ─── ESTRUCTURA DEL DURLOCK ───────────────────────────────────────────────────
# Tareas por piso para el cert de Durlock interno (Julio Cabrera)
DURLOCK_TAREAS = ["Armado de estructura según plano", "Emplacado general", "Masillado e iluminacion"]
DURLOCK_PISOS = ["PLANTA BAJA (LOSA EXISTENTE)", "PISO 1°", "PISO 2°", "PISO 3°", "PISO 4°",
                 "PISO 5°", "PISO 6°", "PISO 7°", "PISO 8°", "PISO 9°", "PISO 10°",
                 "PISO 11°", "PISO 12°", "PISO 13°", "PISO 14°"]
# Valor por piso (presupuesto interno Julio Cabrera)
DURLOCK_VALORES = {
    "PLANTA BAJA (LOSA EXISTENTE)": {"Armado de estructura según plano": 1992879, "Emplacado general": 1992879, "Masillado e iluminacion": 1708182},
    "PISO 1°":  {"Armado de estructura según plano": 2003423, "Emplacado general": 2003423, "Masillado e iluminacion": 1717220},
    "PISO 2°":  {"Armado de estructura según plano": 2003423, "Emplacado general": 2003423, "Masillado e iluminacion": 1717220},
    "PISO 3°":  {"Armado de estructura según plano": 2003423, "Emplacado general": 2003423, "Masillado e iluminacion": 1717220},
    "PISO 4°":  {"Armado de estructura según plano": 2003423, "Emplacado general": 2003423, "Masillado e iluminacion": 1717220},
    "PISO 5°":  {"Armado de estructura según plano": 2003423, "Emplacado general": 2003423, "Masillado e iluminacion": 1717220},
    "PISO 6°":  {"Armado de estructura según plano": 432318, "Emplacado general": 432318, "Masillado e iluminacion": 370558},
    "PISO 7°":  {"Armado de estructura según plano": 432318, "Emplacado general": 432318, "Masillado e iluminacion": 370558},
    "PISO 8°":  {"Armado de estructura según plano": 432318, "Emplacado general": 432318, "Masillado e iluminacion": 370558},
    "PISO 9°":  {"Armado de estructura según plano": 432318, "Emplacado general": 432318, "Masillado e iluminacion": 370558},
    "PISO 10°": {"Armado de estructura según plano": 432318, "Emplacado general": 432318, "Masillado e iluminacion": 370558},
    "PISO 11°": {"Armado de estructura según plano": 405957, "Emplacado general": 405957, "Masillado e iluminacion": 347963},
    "PISO 12°": {"Armado de estructura según plano": 274153, "Emplacado general": 274153, "Masillado e iluminacion": 234988},
    "PISO 13°": {"Armado de estructura según plano": 274153, "Emplacado general": 274153, "Masillado e iluminacion": 234988},
    "PISO 14°": {"Armado de estructura según plano": 274153, "Emplacado general": 274153, "Masillado e iluminacion": 234988},
}
DURLOCK_TOTAL = 44_000_000


def _parse_avances_ia(mensaje, contrato):
    """Usa IA para parsear el avance de pisos desde lenguaje natural."""
    system = f"""Sos un asistente que parsea certificados de construcción para el contrato {contrato}.
El usuario describe avances por piso y tarea. Extraé SOLO los avances mencionados.

Para Durlock, las tareas son: "Armado de estructura según plano", "Emplacado general", "Masillado e iluminacion"
Los pisos son: PLANTA BAJA, PISO 1° a PISO 14°

Respondé SOLO con JSON válido, sin texto extra:
{{
  "fecha": "dd/mm/aaaa o null",
  "incluye_cac": false,
  "monto_cac": 0,
  "avances": {{
    "PISO 8°": {{"Armado de estructura según plano": 10, "Emplacado general": 0, "Masillado e iluminacion": 0}},
    "PISO 9°": {{"Armado de estructura según plano": 10}}
  }}
}}

Si no menciona una tarea en un piso, no la incluyas (se entiende que es 0% actual).
Los porcentajes son el avance ACTUAL (no acumulado)."""

    headers = {"Content-Type": "application/json",
               "x-api-key": ANTHROPIC_API_KEY,
               "anthropic-version": "2023-06-01"}
    body = {"model": "claude-haiku-4-5-20251001", "max_tokens": 500,
            "system": system,
            "messages": [{"role": "user", "content": mensaje}]}
    r = requests.post("https://api.anthropic.com/v1/messages", headers=headers, json=body)
    text = r.json()["content"][0]["text"].strip()
    import re
    text = re.sub(r"^```json|^```|```$", "", text, flags=re.MULTILINE).strip()
    return json.loads(text)


def _leer_acumulados_durlock(wb):
    """Lee los % acumulados del último cert del xlsx de Durlock."""
    acumulados = {}
    for sheet_name in reversed(wb.sheetnames):
        if "CERT" in sheet_name.upper():
            ws = wb[sheet_name]
            piso_actual = None
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        # detectar piso
                        for p in DURLOCK_PISOS:
                            if p in str(cell):
                                piso_actual = p
                        # detectar tarea con % acumulado
                        for t in DURLOCK_TAREAS:
                            if t in str(cell) and piso_actual:
                                # buscar el % ACUMULADO en la misma fila
                                row_list = list(row)
                                idx = next((i for i,v in enumerate(row_list) if v == cell), None)
                                if idx is not None:
                                    # el acumulado suele estar ~4 cols después
                                    for offset in range(2, 8):
                                        if idx+offset < len(row_list):
                                            val = row_list[idx+offset]
                                            if isinstance(val, (int, float)) and 0 <= val <= 1:
                                                key = f"{piso_actual}|{t}"
                                                acumulados[key] = val
                                                break
            break  # solo el último cert
    return acumulados


def _generar_cert_xlsx_durlock(avances_nuevos, fecha_str, num_cert, acumulados_prev):
    """
    Genera el contenido de la nueva hoja CERT N°X para Durlock.
    Retorna (wb modificado, monto_cert, monto_acum_nuevo)
    """
    # Descargar xlsx actual
    xlsx_bytes = drive_download(DRIVE_FILES["durlock"]["xlsx"])
    wb = load_workbook(io.BytesIO(xlsx_bytes))

    # Crear nueva hoja
    sheet_name = f"CERT N°{num_cert}"
    ws = wb.create_sheet(sheet_name)

    # Estilos
    thin = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    def fill(h): return PatternFill("solid", start_color=h, end_color=h)
    F_HDR  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    F_DATA = Font(name="Arial", size=10)
    F_BOLD = Font(name="Arial", bold=True, size=10)
    F_AMB  = Font(name="Arial", size=10, color="CC6600")

    try:
        fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
    except:
        fecha = datetime.now()

    # Header
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = f"CERTIFICADO N°{num_cert}   |   Durlock MO — Santiago del Estero   |   {fecha.strftime('%d/%m/%Y')}"
    c.font = F_HDR; c.fill = fill("1F2D3D"); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Headers de columnas
    headers = ["PISO", "TAREA", "VALOR TOTAL", "ANT. %", "ACTUAL %", "ACUM. %", "ANT. $", "ACTUAL $", "ACUM. $"]
    widths  = [22, 32, 14, 9, 9, 9, 14, 14, 14]
    for i, (h, w) in enumerate(zip(headers, widths)):
        col = i + 1
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = F_HDR; cell.fill = fill("374151")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 32

    # Datos
    row = 3
    monto_cert = 0
    monto_acum = 0
    rows_detalle = []  # para calcular el total al final

    for piso in DURLOCK_PISOS:
        piso_first = True
        for tarea in DURLOCK_TAREAS:
            key = f"{piso}|{tarea}"
            valor_total = DURLOCK_VALORES.get(piso, {}).get(tarea, 0)
            pct_ant = acumulados_prev.get(key, 0.0)
            pct_act = avances_nuevos.get(piso, {}).get(tarea, 0.0) / 100.0
            pct_acum = min(1.0, pct_ant + pct_act)
            imp_ant  = round(valor_total * pct_ant)
            imp_act  = round(valor_total * pct_act)
            imp_acum = round(valor_total * pct_acum)
            monto_cert += imp_act
            monto_acum  += imp_acum

            # Col A: piso (solo primera tarea del piso)
            c_piso = ws.cell(row=row, column=1, value=piso if piso_first else "")
            c_piso.font = F_BOLD if piso_first else F_DATA
            c_piso.alignment = Alignment(vertical="center"); c_piso.border = BORDER
            if piso_first:
                c_piso.fill = fill("E8F4FD")
            piso_first = False

            vals = [tarea, valor_total, pct_ant, pct_act, pct_acum, imp_ant, imp_act, imp_acum]
            for ci, v in enumerate(vals):
                col = ci + 2
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = F_AMB if (pct_act > 0 and col in [5, 8]) else F_DATA
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if col in [3, 7, 8, 9]:
                    cell.number_format = '#,##0;(#,##0);"-"'
                elif col in [4, 5, 6]:
                    cell.number_format = '0.00%;(0.00%);"-"'
            row += 1

    # Totales
    ws.cell(row=row, column=1, value="TOTAL").font = F_BOLD
    ws.cell(row=row, column=3, value=DURLOCK_TOTAL).font = F_BOLD
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=8, value=monto_cert).font = F_BOLD
    ws.cell(row=row, column=8).number_format = '#,##0'
    ws.cell(row=row, column=9, value=monto_acum).font = F_BOLD
    ws.cell(row=row, column=9).number_format = '#,##0'
    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = fill("F0F0F0")
        ws.cell(row=row, column=col).border = BORDER
    row += 1

    # Totales de cert
    ws.cell(row=row+1, column=1, value="MONTO CERT ACTUAL:").font = F_BOLD
    ws.cell(row=row+1, column=3, value=monto_cert).font = F_BOLD
    ws.cell(row=row+1, column=3).number_format = '#,##0'
    ws.cell(row=row+2, column=1, value="TOTAL ACUMULADO:").font = F_BOLD
    ws.cell(row=row+2, column=3, value=monto_acum).font = F_BOLD
    ws.cell(row=row+2, column=3).number_format = '#,##0'
    ws.cell(row=row+3, column=1, value="AVANCE %:").font = F_BOLD
    pct_avance = monto_acum / DURLOCK_TOTAL if DURLOCK_TOTAL else 0
    ws.cell(row=row+3, column=3, value=pct_avance).font = F_BOLD
    ws.cell(row=row+3, column=3).number_format = '0.00%'

    # Guardar
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), monto_cert, monto_acum, pct_avance, wb


def certificar_durlock(mensaje, num_whatsapp=None):
    """
    Función principal que procesa un mensaje de certificación de Durlock.
    Retorna (texto_respuesta, ok: bool)
    """
    try:
        # 1. Parsear avances con IA
        parsed = _parse_avances_ia(mensaje, "Durlock MO")
        avances = parsed.get("avances", {})
        fecha_str = parsed.get("fecha") or datetime.now().strftime("%d/%m/%Y")
        if not avances:
            return "❌ No pude identificar avances. Ejemplo: 'Piso 8 estructura 10%, Piso 9 estructura 15%'", False

        # 2. Descargar xlsx y leer acumulados
        xlsx_bytes = drive_download(DRIVE_FILES["durlock"]["xlsx"])
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        acumulados = _leer_acumulados_durlock(wb)

        # Determinar número de cert
        cert_nums = [int(s.split("N°")[1]) for s in wb.sheetnames if "CERT N°" in s and s.split("N°")[1].isdigit()]
        num_cert = max(cert_nums) + 1 if cert_nums else 1

        # 3. Generar nueva pestaña
        xlsx_nuevo, monto_cert, monto_acum, pct_avance, wb_nuevo = _generar_cert_xlsx_durlock(
            avances, fecha_str, num_cert, acumulados
        )

        # 4. Subir xlsx actualizado al Drive
        drive_update(DRIVE_FILES["durlock"]["xlsx"], xlsx_nuevo,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # 5. Generar PDF y subirlo (opcional - requiere LibreOffice)
        pdf_str = ""
        try:
            pdf_nombre = f"Cs_SDE_DURLOCK_CERTN°{num_cert}.pdf"
            pdf_bytes = _xlsx_a_pdf(xlsx_nuevo, f"CERT N°{num_cert}")
            if pdf_bytes:
                drive_upload_new(DRIVE_FILES["durlock"]["folder"], pdf_nombre, pdf_bytes, "application/pdf")
                pdf_str = " + PDF"
        except Exception as pdf_err:
            print(f"PDF skip: {pdf_err}")

        # 6. Resumen de avances
        lineas_avance = []
        for piso, tareas in avances.items():
            for tarea, pct in tareas.items():
                if pct > 0:
                    tarea_corta = tarea.replace("según plano", "").replace("e iluminacion", "").strip()
                    lineas_avance.append(f"  • {piso} — {tarea_corta}: {pct}%")

        respuesta = (
            f"✅ *Cert N°{num_cert} Durlock registrado*\n"
            f"📅 {fecha_str}\n\n"
            f"*Avances certificados:*\n" + "\n".join(lineas_avance) + "\n\n"
            f"💵 Monto cert: *${monto_cert/1e6:.3f}M*\n"
            f"📊 Acumulado: ${monto_acum/1e6:.3f}M ({pct_avance*100:.1f}% del contrato)\n"
            f"📁 Excel{pdf_str} actualizado en Drive"
        )
        return respuesta, True

    except Exception as e:
        import traceback
        print(f"CERT ERROR: {traceback.format_exc()}")
        return f"❌ Error: {str(e)}", False


def _xlsx_a_pdf(xlsx_bytes, sheet_name):
    """Convierte una hoja de Excel a PDF usando LibreOffice."""
    try:
        import subprocess
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(xlsx_bytes)
            xlsx_path = f.name
        out_dir = tempfile.mkdtemp()
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", out_dir, xlsx_path],
            capture_output=True, timeout=30
        )
        pdf_path = xlsx_path.replace(".xlsx", ".pdf").replace(
            os.path.dirname(xlsx_path), out_dir)
        if os.path.exists(pdf_path):
            with open(pdf_path, "rb") as f:
                return f.read()
    except Exception as e:
        print(f"PDF generation failed: {e}")
    return None
